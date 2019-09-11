"""
========================
@author  :疯疯
@time    :2019/8/23 , 17:39
@E-mail  :1609456360@qq.com
@file    :read_excel.py
@Software: PyCharm
========================
"""
import openpyxl
#测试数据类
class CaseZip(object):
    #用例数据函数
    def cases(self,*args,**kwargs):
        '''用来创建对象属性'''
        case_list = list(*args)  #将传过来的打包的用例数据转换为列表
        for case in case_list:  #遍历列表，添加属性，第一个元素为属性名，第二个元素为值
            setattr(self,case[0],case[1])
#读取Excel表格类
class ReadExcel(object):
    def __init__(self,filename,sheetname):
        '''
        初始化参数
        :param filename: 要打开的文件名
        :param sheetname: 要打开的表单名
        '''
        self.filename = filename
        self.sheetname = sheetname
    def open(self):
        '''打开Excel表单'''
        self.work_book = openpyxl.load_workbook(self.filename)
        self.sheet = self.work_book[self.sheetname]

    def read_excel(self):
        '''读取表格数据'''
        self.open()#调用打开表单方法
        cases = [] #用于接收每个用例数据对象
        rows = list(self.sheet.rows)  #按行获取数据
        #将表头数据存储为一个列表
        title = [row.value for row in rows[0]]
        #遍历除表头的数据
        for row in rows[1:]:
            data = [] #用于存储单条用例数据
            for r in row:
                #每个表格内容存储到列表
                data.append(r.value)
            #将表头和单行内容打包
            case_zip = zip(title, data)
            case_data = CaseZip()
            case_data.cases(case_zip)  #创建用例数据对象
            cases.append(case_data)   #将每个用例对象存到列表中
        return cases  #返回用例对象列表
    def white_excelMax(self,row,value):
        '''
        写入数据到Excel表最后一列
        :param row: 要写入的行数
        :param value: 要写入的内容
        :return:
        '''
        self.open()
        max_column = self.sheet.max_column  #获取最大列数
        cell = self.sheet.cell(row = row,column = max_column)  #获取要写入的表格
        cell.value = value    #写入内容
        self.work_book.save(self.filename)  #保存数据
    def white_excel(self,row,value,column):
        '''
        写入数据到Excel表指定列
        :param row: 要写入的行数
        :param value: 要写入的内容
        :return:
        '''
        self.open()
        cell = self.sheet.cell(row = row,column = column)  #获取要写入的表格
        cell.value = value    #写入内容
        self.work_book.save(self.filename)  #保存数据